import argparse
import glob
import json
import csv
import re
from openpyxl import load_workbook
import xml.etree.ElementTree as ET

'''
This script will read in one file at a time, and transform into a dbGaP XML.
Loops will be done outside the script, e.g. in a bash script that can be executed
over multiple files.
Option to keep the JSON intermediate.
'''

def main(args): 
    '''
    colname_transforms - transforms the colnames in the source data dictionaries
    to the desired name in the final output.  Assumes homogeneity among all data
    dictionaries that are being transformed. It is a dictionary where the key is
    current name and value is desired name.

    '''
    colname_transforms = {}
    if args.var_id_column:
        colname_transforms['variable_id'] = args.var_id_column
    if args.var_name_column:
        colname_transforms['variable_name'] = args.var_name_column
    if args.var_desc_column:
        colname_transforms['variable_description'] = args.var_desc_column

    # Transform to JSON
    json_blob = nida_raw_to_json(data_dictionary = args.dd, study_id = args.study_id, colname_transforms = colname_transforms)

    if args.json_path:
        filename = f"{args.json_path}/{args.study_id}.json"
        with open(filename, "w") as stream:
            json.dump(json_blob, stream)

    for dataset in json_blob:
        # Transform to dbGaP XML
        dataset_xml = json_to_dbgap_xml(dataset)
        # Format XML
        ET.indent(dataset_xml)
        # Write xml
        filename = f"{args.output_path}/{dataset['study_id']}.{dataset['dataset_id']}.xml"
        dataset_xml.write(filename)

def nida_raw_to_json(data_dictionary, study_id, colname_transforms):
    '''
    Transforms the raw NIDA data for a single study ID (multiple datasets) and 
    returns single JSON blob of roughly the following format:

    [
        {
            "study_id": ,
            "dataset_id": ,
            "dataset_name": ,
            "variables": [
                {
                    "variable_id": ,
                    "variable_name": ,
                    "variable_description":
                },
            ]
        },
        ...
    ]
    '''
    json_blob = [] # initialize output json blob
   
    wb = load_workbook(data_dictionary) 
    for sheet in wb.sheetnames:
        dataset = {} # initialize dataset
        dataset['study_id'] = study_id

        # dataset metadata
        dataset_id = re.sub("[^a-z ]","",sheet.lower()).replace(" ","_")
        dataset['dataset_id'] = dataset_id
        dataset['dataset_name'] = sheet
        dataset['variables'] = []
        
        # variable metadata
        rows = wb[sheet].rows
        
        # filter keys - only keep keys in colname_transforms
        keys = [k.internal_value for k in next(rows)]
        for iteration,row in enumerate(rows):

            # Get values for each dictionary
            values = [k.internal_value for k in row]

            # Skip empty rows
            if values[0] is None:
                continue

            variable = {k:v for k,v in zip(keys, values) if k in [val for vals in colname_transforms.values() for val in vals]}
            new_variable = {}

            # Change key name
            for v in variable:
                new_key = [k for k in colname_transforms if v in colname_transforms[k]].pop()
                new_variable[new_key] = variable[v]

            # variable id missing?
            if not 'variable_id' in new_variable:
                new_variable['variable_id'] = f"{dataset_id}.v{iteration+1}"
            
            # Append variable
            dataset['variables'].append(new_variable)

        # Append dataset
        json_blob.append(dataset)
    
    return(json_blob)

def json_to_dbgap_xml(dataset):
    '''
    Transforms the JSON blob into a dbGaP XML format, e.g.:

    <data_table id="pht000700.v1" study_id="phs000166.v2" participant_set="1" 
    date_created="Thu Sep  3 15:21:50 2009">

    <variable id="phv00070931.v1">
        <name>SUBJ_ID</name>
        <description>Deidentified Subject ID</description>
        <type>integer</type>
    </variable>
    '''
    # Build root
    root = ET.Element("data_table")
    root.set("id",dataset["dataset_id"])
    root.set("study_id",dataset["study_id"])

    # Loop over each variable
    for var_dict in dataset['variables']:
        variable = ET.SubElement(root,"variable")
        variable.set("id",var_dict['variable_id'])
        name = ET.SubElement(variable, "name")
        name.text = var_dict.get('variable_name',"")
        desc = ET.SubElement(variable, "description")
        desc.text = var_dict.get('variable_description',"")

    return(ET.ElementTree(root))

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Transform NIDA to JSON and/or dbGaP XML format")
    parser.add_argument('dd', action="store", help= "Specify the file to convert")
    parser.add_argument('study_id', action="store", help = "Specify the study ID")
    parser.add_argument('output_path', action="store", help ="Specify absolute path for outputs")
    parser.add_argument('-j', '--json', dest="json_path", action="store", help = "Specify JSON output path (optional)" )

    # add column names
    parser.add_argument('--var-id-column', nargs="+", dest="var_id_column", action="store", help = "Specify the column(s) in the file which contains the variable ID")
    parser.add_argument('--var-name-column', nargs="+", dest="var_name_column", action="store", help = "Specify the column(s) in the file which contains the variable name")
    parser.add_argument('--var-desc-column', nargs="+", dest="var_desc_column", action="store", help = "Specify the column(s) in the file which contains the variable description")

    args = parser.parse_args()
    main(args)